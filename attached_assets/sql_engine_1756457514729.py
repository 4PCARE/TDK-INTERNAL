import csv
import os
import re
import sqlite3
import threading
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
import dateutil.parser

try:
    import openpyxl  # optional for .xlsx
except Exception:
    openpyxl = None

class SimpleSQLite:
    def __init__(self, table_name: str):
        self.table_name = table_name
        self.conn = sqlite3.connect(":memory:", check_same_thread=False)
        self.lock = threading.Lock()
        self.conn.row_factory = sqlite3.Row

    def _infer_type(self, s: str) -> str:
        if s is None:
            return "TEXT"
        s = str(s).strip()
        if not s:
            return "TEXT"
            
        # int?
        if re.fullmatch(r"[+-]?\d+", s):
            return "INTEGER"
        # float?
        if re.fullmatch(r"[+-]?(\d*\.\d+|\d+\.\d*)([eE][+-]?\d+)?", s):
            return "REAL"
        # datetime detection - check various formats
        if self._is_datetime(s):
            return "DATETIME"
        return "TEXT"
    
    def _is_datetime(self, s: str) -> bool:
        """Check if string represents a datetime value"""
        try:
            # Common datetime patterns
            datetime_patterns = [
                r"\d{4}-\d{2}-\d{2}",  # YYYY-MM-DD
                r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}",  # YYYY-MM-DD HH:MM:SS
                r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}",  # ISO format
                r"\d{2}/\d{2}/\d{4}",  # MM/DD/YYYY
                r"\d{2}-\d{2}-\d{4}",  # MM-DD-YYYY
                r"\d{1,2}/\d{1,2}/\d{4}",  # M/D/YYYY
            ]
            
            # Check patterns first for performance
            for pattern in datetime_patterns:
                if re.fullmatch(pattern, s):
                    # Try to actually parse it to confirm
                    try:
                        dateutil.parser.parse(s)
                        return True
                    except:
                        continue
            
            # Try parsing with dateutil for other formats
            try:
                dateutil.parser.parse(s)
                return True
            except:
                return False
        except:
            return False
    
    def _parse_datetime(self, s: str) -> str:
        """Parse datetime string to ISO format for SQLite"""
        if s is None or str(s).strip() == "":
            return None
        try:
            parsed = dateutil.parser.parse(str(s))
            return parsed.isoformat()
        except:
            return str(s)  # Return original if parsing fails

    def load_rows(self, rows: List[Dict[str, Any]]):
        if not rows:
            raise ValueError("No rows to load")
        # infer types from first non-empty row
        cols = list(rows[0].keys())
        types = []
        for c in cols:
            v = next((r.get(c) for r in rows if r.get(c) not in (None, "")), None)
            types.append(self._infer_type(v) if v is not None else "TEXT")

        # Map DATETIME to TEXT for SQLite storage but track datetime columns
        datetime_cols = {c for c, t in zip(cols, types) if t == "DATETIME"}
        numeric_cols = {c for c, t in zip(cols, types) if t in ("INTEGER", "REAL")}
        sql_types = ["TEXT" if t == "DATETIME" else t for t in types]
        
        col_defs = ", ".join(f'"{c}" {t}' for c, t in zip(cols, sql_types))
        with self.lock:
            self.conn.execute(f'DROP TABLE IF EXISTS "{self.table_name}"')
            self.conn.execute(f'CREATE TABLE "{self.table_name}" ({col_defs})')

            placeholders = ", ".join("?" for _ in cols)
            insert_sql = f'INSERT INTO "{self.table_name}" ({", ".join([f'"{c}"' for c in cols])}) VALUES ({placeholders})'
            for r in rows:
                values = []
                for c in cols:
                    val = r.get(c, None)
                    # Convert empty strings to None for numeric columns
                    if c in numeric_cols and val == "":
                        val = None
                    # Parse datetime values to ISO format
                    elif c in datetime_cols and val is not None and val != "":
                        val = self._parse_datetime(val)
                    values.append(val)
                self.conn.execute(insert_sql, values)
            self.conn.commit()

    def execute_safe_select(self, sql: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        # Clean and normalize the SQL
        original_sql = sql.strip()
        s = original_sql
        
        # Remove common prefixes that might be added by LLMs
        if s.startswith("```sql"):
            s = s[6:]
        if s.startswith("```"):
            s = s[3:]
        if s.endswith("```"):
            s = s[:-3]
        
        s = s.strip()
        
        # Allow SELECT, WITH, and other read-only operations
        allowed_starts = [
            "SELECT", "select", "WITH", "with", 
            "SHOW", "show", "DESCRIBE", "describe", "DESC", "desc", "EXPLAIN", "explain",
            "PRAGMA", "pragma"
        ]
        
        # Find the first meaningful SQL statement (case-insensitive but preserve original case)
        lines = s.split('\n')
        first_meaningful_line = ""
        for line in lines:
            clean_line = line.strip()
            if clean_line and not clean_line.startswith('--') and not clean_line.startswith('/*'):
                first_meaningful_line = clean_line
                break
        
        # If no meaningful line found, use the whole statement
        if not first_meaningful_line:
            first_meaningful_line = s.strip()
        
        # Check if this looks like explanatory text rather than SQL
        explanatory_markers = ["post_filter:", "no matching", "reason:", "no sql needed"]
        if any(text in first_meaningful_line.lower() for text in explanatory_markers):
            raise ValueError(f"Expected SQL query, but got explanatory text: {first_meaningful_line[:100]}...")
        
        # Check if it's a valid read-only query (case-insensitive)
        first_word = first_meaningful_line.split()[0].upper() if first_meaningful_line.split() else ""
        allowed_first_words = ["SELECT", "WITH", "SHOW", "DESCRIBE", "DESC", "EXPLAIN", "PRAGMA"]
        
        if first_word not in allowed_first_words:
            raise ValueError(f"Only SELECT/WITH and other read-only queries are allowed in demo mode. Got: {first_word} (from: {first_meaningful_line[:50]}...)")
        
        # Execute the original SQL (not the processed version)
        with self.lock:
            cur = self.conn.execute(original_sql)
            col_names = [d[0] for d in cur.description]
            data = [dict(row) for row in cur.fetchall()]
        return col_names, data

    def schema_text(self, sample_rows: int = 3) -> str:
        # build CREATE TABLE-ish schema description
        with self.lock:
            info = self.conn.execute(f'PRAGMA table_info("{self.table_name}")').fetchall()
            cols = [(row[1], row[2]) for row in info]  # name, type
            sample = self.conn.execute(f'SELECT * FROM "{self.table_name}" LIMIT {sample_rows}').fetchall()
        
        # Detect datetime columns by examining sample data
        datetime_cols = set()
        if sample:
            for row in sample:
                for col_name, _ in cols:
                    val = row[col_name] if col_name in row.keys() else None
                    if val and self._is_datetime(str(val)):
                        datetime_cols.add(col_name)
        
        lines = [f'CREATE TABLE "{self.table_name}" (']
        for name, typ in cols:
            # Show DATETIME type for datetime columns in schema
            display_type = "DATETIME" if name in datetime_cols else typ
            lines.append(f'  "{name}" {display_type},')
        lines[-1] = lines[-1].rstrip(",")
        lines.append(");")
        
        # Add datetime handling notes
        if datetime_cols:
            lines.append(f"\n-- DATETIME COLUMNS: {', '.join(sorted(datetime_cols))}")
            lines.append("-- These columns contain parsed datetime values in ISO format")
            lines.append("-- Use DATETIME() function for comparisons: DATETIME(column_name) >= DATETIME('now')")
        
        # Add a note about quoting column names with special characters
        lines.append("\n-- NOTE: Column names with special characters (parentheses, spaces, etc.) must be quoted with double quotes")
        lines.append(f'-- Example: SELECT "multiple_promotion_eligibility_(extra_flag)" FROM "{self.table_name}";')
        
        if sample:
            lines.append("\n-- Sample rows:")
            # show a couple of rows with quoted column names in the representation
            for r in sample:
                as_dict = {f'"{k}"': r[k] for k in r.keys()}
                lines.append(str(as_dict).replace("'\"", '"').replace("\"'", '"'))
        return "\n".join(lines)

def load_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)

def load_xlsx(path: str) -> List[Dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed; cannot read .xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        out.append({h: v for h, v in zip(headers, r)})
    return out

def build_engine_from_file(path: str, table_name: str) -> SimpleSQLite:
    _, ext = os.path.splitext(path.lower())
    if ext == ".csv":
        rows = load_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        rows = load_xlsx(path)
    else:
        raise ValueError("Unsupported file type; use .csv or .xlsx")
    engine = SimpleSQLite(table_name=table_name)
    engine.load_rows(rows)
    return engine
